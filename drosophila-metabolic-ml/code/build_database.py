"""
build_database.py
------------------
Parses the Drosophila genome-scale metabolic model (Reaction List +
Metabolite List sheets) and loads it into a normalized SQLite database.

Input : data/Drosophila.xlsx
Output: data/metabolic_model.db

Tables created:
    metabolites         (abbreviation, description, formula, compartment)
    reactions           (abbreviation, description, equation, lower_bound,
                          upper_bound, is_objective, subsystem, is_reversible,
                          is_exchange, is_transport)
    genes               (fly_base_id)
    reaction_genes      (reaction_abbr, gene_id)   -- many-to-many, GPR exploded
    reaction_metabolites(reaction_abbr, metabolite_abbr, stoichiometry, role)
                          role = 'reactant' or 'product'
"""

import re
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
DATA_DIR = HERE.parent / "data"
XLSX_PATH = DATA_DIR / "Drosophila.xlsx"
DB_PATH = DATA_DIR / "metabolic_model.db"

# Matches a term like "0.26551 m03161[c]" or "m02039[c]" (coeff optional)
TERM_RE = re.compile(r"^\s*([0-9.eE+\-]+)?\s*(m\d+\[[a-z]\])\s*$")
GENE_TOKEN_RE = re.compile(r"FBgn\d+")


def parse_equation(equation: str):
    """Split a reaction equation string into (reactants, products, reversible).

    Reactants/products are lists of (metabolite_abbr, stoichiometric_coeff).
    """
    if "<=>" in equation:
        reversible = True
        lhs, rhs = equation.split("<=>")
    elif "->" in equation:
        reversible = False
        lhs, rhs = equation.split("->")
    else:
        # Exchange reactions with nothing on one side still contain one arrow;
        # if neither found, treat the whole thing as malformed and skip.
        return [], [], False

    def parse_side(side: str):
        terms = []
        for chunk in side.split("+"):
            chunk = chunk.strip()
            if not chunk:
                continue
            m = TERM_RE.match(chunk)
            if not m:
                continue
            coeff_str, met = m.groups()
            coeff = float(coeff_str) if coeff_str else 1.0
            terms.append((met, coeff))
        return terms

    reactants = parse_side(lhs)
    products = parse_side(rhs)
    return reactants, products, reversible


def parse_gpr(gpr: str):
    """Extract the set of FlyBase gene IDs referenced in a GPR boolean string."""
    if not gpr:
        return []
    return sorted(set(GENE_TOKEN_RE.findall(gpr)))


def build_schema(conn: sqlite3.Connection):
    conn.executescript(
        """
        DROP TABLE IF EXISTS metabolites;
        DROP TABLE IF EXISTS reactions;
        DROP TABLE IF EXISTS genes;
        DROP TABLE IF EXISTS reaction_genes;
        DROP TABLE IF EXISTS reaction_metabolites;

        CREATE TABLE metabolites (
            abbreviation TEXT PRIMARY KEY,
            description  TEXT,
            formula      TEXT,
            compartment  TEXT
        );

        CREATE TABLE reactions (
            abbreviation   TEXT PRIMARY KEY,
            description    TEXT,
            equation       TEXT,
            lower_bound    REAL,
            upper_bound    REAL,
            is_objective   INTEGER,
            subsystem      TEXT,
            is_reversible  INTEGER,
            is_exchange    INTEGER,
            is_transport   INTEGER
        );

        CREATE TABLE genes (
            fly_base_id TEXT PRIMARY KEY
        );

        CREATE TABLE reaction_genes (
            reaction_abbr TEXT,
            gene_id       TEXT,
            FOREIGN KEY (reaction_abbr) REFERENCES reactions(abbreviation),
            FOREIGN KEY (gene_id) REFERENCES genes(fly_base_id)
        );

        CREATE TABLE reaction_metabolites (
            reaction_abbr   TEXT,
            metabolite_abbr TEXT,
            stoichiometry   REAL,
            role            TEXT,
            FOREIGN KEY (reaction_abbr) REFERENCES reactions(abbreviation),
            FOREIGN KEY (metabolite_abbr) REFERENCES metabolites(abbreviation)
        );

        CREATE INDEX idx_rg_reaction ON reaction_genes(reaction_abbr);
        CREATE INDEX idx_rg_gene ON reaction_genes(gene_id);
        CREATE INDEX idx_rm_reaction ON reaction_metabolites(reaction_abbr);
        CREATE INDEX idx_rm_metabolite ON reaction_metabolites(metabolite_abbr);
        """
    )


def load_metabolites(conn, ws):
    rows = ws.iter_rows(min_row=2, values_only=True)
    data = []
    for r in rows:
        if not r or not r[0]:
            continue
        abbr, desc, formula, compartment = (r + (None,) * 4)[:4]
        data.append((abbr, desc, formula, compartment))
    conn.executemany(
        "INSERT OR IGNORE INTO metabolites VALUES (?,?,?,?)", data
    )
    return len(data)


def load_reactions(conn, ws):
    rows = ws.iter_rows(min_row=2, values_only=True)

    reaction_rows = []
    gene_set = set()
    reaction_gene_rows = []
    reaction_met_rows = []

    for r in rows:
        if not r or not r[0]:
            continue
        abbr, desc, equation, gpr, lb, ub, obj, subsystem = (r + (None,) * 8)[:8]

        lb = float(lb) if lb not in (None, "") else -1000.0
        ub = float(ub) if ub not in (None, "") else 1000.0
        is_objective = 1 if str(obj) not in ("0", "0.0", "", "None") else 0

        reactants, products, reversible = parse_equation(equation or "")
        is_exchange = 1 if (len(reactants) == 0 or len(products) == 0) else 0

        # Transport: same metabolite (ignoring compartment tag) appears on
        # both sides, or metabolites span more than one compartment.
        def base_id(m):
            return m.split("[")[0]

        compartments = set()
        for m, _ in reactants + products:
            compartments.add(m.split("[")[-1].rstrip("]"))
        is_transport = 1 if len(compartments) > 1 else 0

        reaction_rows.append(
            (
                abbr,
                desc,
                equation,
                lb,
                ub,
                is_objective,
                subsystem,
                1 if reversible else 0,
                is_exchange,
                is_transport,
            )
        )

        for met, coeff in reactants:
            reaction_met_rows.append((abbr, met, -abs(coeff), "reactant"))
        for met, coeff in products:
            reaction_met_rows.append((abbr, met, abs(coeff), "product"))

        for gene in parse_gpr(gpr or ""):
            gene_set.add(gene)
            reaction_gene_rows.append((abbr, gene))

    conn.executemany(
        "INSERT OR IGNORE INTO reactions VALUES (?,?,?,?,?,?,?,?,?,?)",
        reaction_rows,
    )
    conn.executemany(
        "INSERT OR IGNORE INTO genes VALUES (?)", [(g,) for g in gene_set]
    )
    conn.executemany(
        "INSERT INTO reaction_genes VALUES (?,?)", reaction_gene_rows
    )
    conn.executemany(
        "INSERT INTO reaction_metabolites VALUES (?,?,?,?)", reaction_met_rows
    )

    return len(reaction_rows), len(gene_set)


def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"Expected converted workbook at {XLSX_PATH}. "
            "Run the .xls -> .xlsx conversion first."
        )

    wb = load_workbook(XLSX_PATH, read_only=True)
    conn = sqlite3.connect(DB_PATH)

    build_schema(conn)
    n_mets = load_metabolites(conn, wb["Metabolite List"])
    n_rxns, n_genes = load_reactions(conn, wb["Reaction List"])
    conn.commit()

    print(f"Loaded {n_mets} metabolites")
    print(f"Loaded {n_rxns} reactions")
    print(f"Loaded {n_genes} unique genes")
    print(f"Database written to {DB_PATH}")

    conn.close()


if __name__ == "__main__":
    main()
